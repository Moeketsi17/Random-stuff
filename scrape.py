"""
Cylex.net.za business directory scraper.

Two stages:
  1. parse_search_results() - pulls the listing cards from a search/category page
     (name, phone, address, short description, detail page URL, company id)
  2. parse_company_page()   - visits each detail page and pulls the full record
     (address, phone/mobile/fax, website, email, social links, company id)

Fetching is done with a real (visible) Chromium browser via Playwright,
since the site sits behind Cloudflare and returns a JS challenge page to
plain HTTP clients like `requests`.

Usage:
    python scrape.py --query accountants --city Pretoria --pages 3 --out results.xlsx

Be a good citizen: this script rate-limits requests and sends a normal
browser User-Agent. Check https://www.cylex.net.za/robots.txt and their
terms of service before scraping at scale, and keep request volume low.
"""

import argparse
import re
import time
import sys
from urllib.parse import urljoin, urlencode

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright

# Preferred column order for the output sheet; any unexpected extra fields
# are appended at the end so nothing is silently dropped.
COLUMN_ORDER = [
    "company_id", "name", "phone", "phones", "mobiles", "faxes",
    "address", "latlng", "website", "website_redirect_link", "email",
    "social_links", "description", "detail_url", "source_url",
]

BASE = "https://www.cylex.net.za"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)

SOCIAL_DOMAINS = [
    "facebook.com", "m.me", "instagram.com", "twitter.com", "x.com",
    "linkedin.com", "youtube.com", "tiktok.com", "wa.me", "whatsapp.com",
]

# Module-level Playwright/browser handle, set up by start_browser() /
# torn down by stop_browser(). The browser process is reused across
# requests, but each get() opens a brand-new context (= fresh cookie jar)
# and closes it afterwards. The site's WAF blocks a 2nd request that reuses
# a session's cookies in quick succession (e.g. search page -> detail page),
# even from a real browser -- but lets a fresh, cookie-less session through
# every time, so a new context per request is what actually gets served.
_PLAYWRIGHT = None
_BROWSER = None


def start_browser(headless=False):
    """Launch a real Chromium browser so Cloudflare's JS challenge runs and
    clears normally. headless=False by default because Cloudflare is more
    likely to challenge/block headless browsers."""
    global _PLAYWRIGHT, _BROWSER
    _PLAYWRIGHT = sync_playwright().start()
    _BROWSER = _PLAYWRIGHT.chromium.launch(headless=headless)


def stop_browser():
    global _PLAYWRIGHT, _BROWSER
    if _BROWSER:
        _BROWSER.close()
    if _PLAYWRIGHT:
        _PLAYWRIGHT.stop()
    _PLAYWRIGHT = _BROWSER = None


def get(url, params=None, retries=3, delay=1.5):
    """Fetch a page in its own fresh browser context (own cookie jar) with
    basic retry + polite delay, return its HTML."""
    full_url = f"{url}?{urlencode(params)}" if params else url
    for attempt in range(retries):
        context = _BROWSER.new_context(user_agent=USER_AGENT, locale="en-ZA")
        try:
            page = context.new_page()
            page.goto(full_url, timeout=30000)
            time.sleep(delay)
            return page.content()
        except Exception as e:
            print(f"  retry {attempt+1}/{retries} for {url}: {e}", file=sys.stderr)
            time.sleep(delay * (attempt + 1))
        finally:
            context.close()
    return None


def extract_company_id(url_or_divid):
    """Pull the numeric Cylex company id from a URL or a 'fnr12345678' div id."""
    m = re.search(r"(\d{5,})", url_or_divid or "")
    return m.group(1) if m else None


def parse_search_results(html):
    """Parse a search/category listing page -> list of dicts (summary rows)."""
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    # Each listing is a div like: <div id="fnr23792353" class="lm-comp ...">
    for card in soup.select("div[id^='fnr']"):
        name_tag = card.select_one(".h4 a")
        name = name_tag.get_text(strip=True) if name_tag else None
        detail_url = urljoin(BASE, name_tag["href"]) if name_tag and name_tag.get("href") else None

        addr_tag = card.select_one(".addr")
        address = addr_tag.get_text(strip=True) if addr_tag else None

        phone_tag = card.select_one(".lm-ph span")
        phone = phone_tag.get_text(strip=True) if phone_tag else None

        desc_tag = card.select_one(".ellipsis-two-line")
        description = desc_tag.get_text(strip=True) if desc_tag else None

        # lat/lng sometimes sit in a hidden <label> right after the opening row
        latlng_tag = card.find_previous_sibling("label") or card.select_one("label.d-none")
        latlng = latlng_tag.get_text(strip=True) if latlng_tag else None

        rows.append({
            "company_id": extract_company_id(card.get("id")),
            "name": name,
            "phone": phone,
            "address": address,
            "description": description,
            "latlng": latlng,
            "detail_url": detail_url,
        })
    return rows


def parse_company_page(html, url=None):
    """Parse an individual company detail page -> single dict (full record)."""
    soup = BeautifulSoup(html, "html.parser")

    name_tag = soup.select_one("#cp-name span")
    name = name_tag.get_text(strip=True) if name_tag else None

    street_tag = soup.select_one("#cp-street")
    address = ", ".join(s.get_text(strip=True) for s in street_tag.select("span")) if street_tag else None

    # Phones/mobiles/faxes all live in rows with id like cp-mainPhone-1, cp-mainPhone-2, etc,
    # plus extra ones hidden under #cp-morePhones. Distinguish by which icon is used.
    phones, mobiles, faxes = [], [], []
    for row in soup.select("[id^='cp-mainPhone'], #cp-morePhones .row"):
        icon = row.select_one("img")
        number_tag = row.select_one("a[href^='tel:']")
        if not number_tag:
            continue
        number = number_tag.get_text(strip=True)
        icon_src = (icon.get("src") or "") if icon else ""
        if "mobile" in icon_src:
            mobiles.append(number)
        elif "fax" in icon_src:
            faxes.append(number)
        else:
            phones.append(number)

    website_tag = soup.select_one("#cp-website a")
    website = website_tag.get_text(strip=True) if website_tag else None
    website_redirect = website_tag["href"] if website_tag and website_tag.get("href") else None

    email_tag = soup.select_one("#cp-email a[href^='mailto:']")
    email = email_tag["href"].replace("mailto:", "").strip() if email_tag else None

    # Social / messaging links: scan all <a> tags for known social domains
    social_links = {}
    for a in soup.select("a[href]"):
        href = a["href"]
        for domain in SOCIAL_DOMAINS:
            if domain in href:
                social_links.setdefault(domain, href)

    # Company id: prefer the URL, fall back to any onclick handler with the numeric id
    company_id = extract_company_id(url)
    if not company_id:
        edit_div = soup.select_one("[onclick*='OpenUpdateCompanyInformationModal']")
        if edit_div:
            m = re.search(r"OpenUpdateCompanyInformationModal\([^,]+,[^,]+,[^,]+,(\d+)", edit_div["onclick"])
            company_id = m.group(1) if m else None

    return {
        "company_id": company_id,
        "name": name,
        "address": address,
        "phones": "; ".join(phones),
        "mobiles": "; ".join(mobiles),
        "faxes": "; ".join(faxes),
        "website": website,
        "website_redirect_link": website_redirect,
        "email": email,
        "social_links": social_links,  # dict, flattened later for CSV
        "source_url": url,
    }


def search_urls(query, city, pages=1):
    """Yield search-results page URLs for the given query/city across N pages."""
    for p in range(1, pages + 1):
        yield f"{BASE}/s", {
            "q": query, "c": city, "z": "0", "p": str(p),
            "dst": "", "sUrl": "", "cUrl": city.lower(), "he": "1",
        }


def flatten_social(links_dict):
    return "; ".join(f"{k}:{v}" for k, v in links_dict.items())


def run(query, city, pages, out_path, fetch_details=True, delay=1.5, headless=False):
    start_browser(headless=headless)
    try:
        all_rows = []
        for url, params in search_urls(query, city, pages):
            print(f"Fetching search page {params['p']}: {url}?q={query}&c={city}")
            html = get(url, params=params, delay=delay)
            if not html:
                continue
            rows = parse_search_results(html)
            print(f"  found {len(rows)} listings")

            if fetch_details:
                for row in rows:
                    if not row["detail_url"]:
                        continue
                    print(f"    -> {row['name']}")
                    detail_html = get(row["detail_url"], delay=delay)
                    if not detail_html:
                        continue
                    detail = parse_company_page(detail_html, url=row["detail_url"])
                    merged = {**row, **detail}
                    merged["social_links"] = flatten_social(detail["social_links"])
                    all_rows.append(merged)
            else:
                for row in rows:
                    row["social_links"] = ""
                    all_rows.append(row)
    finally:
        stop_browser()

    if not all_rows:
        print("No results scraped.")
        return

    all_keys = {k for row in all_rows for k in row.keys()}
    fieldnames = [c for c in COLUMN_ORDER if c in all_keys]
    fieldnames += sorted(all_keys - set(fieldnames))
    save_xlsx(all_rows, fieldnames, out_path)
    print(f"\nSaved {len(all_rows)} records to {out_path}")


def save_xlsx(rows, fieldnames, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])

    widths = {f: max(len(f), 12) for f in fieldnames}
    for row in rows:
        for f in fieldnames:
            widths[f] = min(max(widths[f], len(str(row.get(f, "")))), 60)
    for i, f in enumerate(fieldnames, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths[f] + 2

    wb.save(out_path)


if __name__ == "__main__":
    # ------------------------------------------------------------------
    # Set your search directly here instead of passing command-line flags.
    # ------------------------------------------------------------------
    QUERY = "accountants"
    CITY = "Pretoria"
    PAGES = 3
    OUT_FILE = "cylex_results.xlsx"
    FETCH_DETAILS = True   # False = skip company pages (faster, less data)
    DELAY = 1.5            # seconds between requests

    # Command-line flags still work and override the values above if given.
    parser = argparse.ArgumentParser(description="Scrape Cylex.net.za business listings")
    parser.add_argument("--query", default=QUERY, help="e.g. accountants")
    parser.add_argument("--city", default=CITY, help="e.g. Pretoria")
    parser.add_argument("--pages", type=int, default=PAGES, help="number of search result pages")
    parser.add_argument("--out", default=OUT_FILE)
    parser.add_argument("--no-details", action="store_true", default=not FETCH_DETAILS,
                         help="skip visiting each company page (faster, less data)")
    parser.add_argument("--delay", type=float, default=DELAY, help="seconds between requests")
    parser.add_argument("--headless", action="store_true",
                         help="run the browser headless (more likely to be challenged/blocked)")
    args = parser.parse_args()

    run(args.query, args.city, args.pages, args.out,
        fetch_details=not args.no_details, delay=args.delay, headless=args.headless)